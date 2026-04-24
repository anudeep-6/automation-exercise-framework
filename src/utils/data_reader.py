"""Reads test data from CSV and JSON files in the test_data directory."""

import csv
import json
import logging
import os
from typing import Optional

from src.utils.exceptions import TestDataException

logger = logging.getLogger(__name__)


class DataReader:
    """Reads test data from CSV and JSON files."""

    def __init__(self, data_dir: str = "test_data") -> None:
        """Initialises DataReader with the test data directory.

        Args:
            data_dir (str): Path to the test data directory.
                Defaults to 'test_data'.
        """
        self.data_dir = data_dir
        logger.info("[DATA] DataReader initialised — data_dir: %s", data_dir)

    def _get_file_path(self, file_name: str) -> str:
        """Builds and validates the full file path.

        Args:
            file_name (str): Name of the file to read.

        Returns:
            str: Full path to the file.

        Raises:
            TestDataException: If the file does not exist at the resolved path.
        """
        file_path = os.path.join(self.data_dir, file_name)
        if not os.path.exists(file_path):
            logger.error("[DATA] File not found: %s", file_path)
            raise TestDataException(file_path, "file not found")
        return file_path

    def read_csv(self, file_name: str) -> list[dict]:
        """Reads a CSV file and returns a list of row dictionaries.

        Each row becomes a dictionary where keys are column headers.

        Args:
            file_name (str): Name of the CSV file.

        Returns:
            list[dict]: List of dictionaries, one per row.

        Raises:
            TestDataException: If the file is malformed or unreadable.
        """
        file_path = self._get_file_path(file_name)
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                reader = csv.DictReader(file)
                rows = list(reader)
            if not rows:
                logger.warning("[DATA] CSV file is empty (no data rows): %s", file_name)
            else:
                logger.info("[DATA] CSV loaded — %s: %d rows", file_name, len(rows))
            return rows
        except csv.Error as err:
            logger.error("[DATA] Failed to parse CSV %s: %s", file_name, err)
            raise TestDataException(file_name, str(err))

    def read_json(self, file_name: str) -> dict | list:
        """Reads a JSON file and returns parsed data.

        Args:
            file_name (str): Name of the JSON file.

        Returns:
            dict | list: Parsed JSON data.

        Raises:
            TestDataException: If the file contains invalid JSON.
        """
        file_path = self._get_file_path(file_name)
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                data = json.load(file)
            logger.info("[DATA] JSON loaded — %s", file_name)
            return data
        except json.JSONDecodeError as err:
            logger.error("[DATA] Failed to parse JSON %s: %s", file_name, err)
            raise TestDataException(file_name, str(err))

    def load_csv_rows(
        self, file_name: str, filter_by: Optional[dict] = None
    ) -> list[dict]:
        """Loads CSV rows with optional column-based filtering.

        A convenience wrapper over read_csv() that filters rows by one or more
        column values without requiring manual list comprehension in test files.

        Args:
            file_name (str): Name of the CSV file to read.
            filter_by (dict, optional): Column-value pairs to filter rows by.
                Only rows matching ALL specified key-value pairs are returned.
                Defaults to None, which returns all rows unfiltered.

        Returns:
            list[dict]: Filtered list of row dictionaries.

        Raises:
            TestDataException: If the file does not exist or is malformed.
        """
        rows = self.read_csv(file_name)
        if filter_by:
            rows = [
                row
                for row in rows
                if all(row.get(k) == v for k, v in filter_by.items())
            ]
            logger.info(
                "[DATA] load_csv_rows — %s: %d rows after filter %s",
                file_name,
                len(rows),
                filter_by,
            )
        return rows

    def load_json_rows(
        self, file_name: str, filter_by: Optional[dict] = None
    ) -> list[dict]:
        """Loads JSON rows with optional key-based filtering.

        Mirrors the load_csv_rows() contract so callers can swap data sources
        without changing filtering logic. The JSON file must contain a
        top-level list of objects.

        Args:
            file_name (str): Name of the JSON file to read.
            filter_by (dict, optional): Key-value pairs to filter rows by.
                Only rows matching ALL specified pairs are returned.
                Defaults to None, which returns all rows unfiltered.

        Returns:
            list[dict]: Filtered list of row dictionaries.

        Raises:
            TestDataException: If the file does not exist, is malformed, or
                does not contain a top-level list.
        """
        data = self.read_json(file_name)
        if not isinstance(data, list):
            logger.error(
                "[DATA] JSON file %s must contain a top-level list, got %s",
                file_name,
                type(data).__name__,
            )
            raise TestDataException(file_name, "expected a top-level JSON array")

        rows = data
        if filter_by:
            rows = [
                row
                for row in rows
                if all(row.get(k) == v for k, v in filter_by.items())
            ]
            logger.info(
                "[DATA] load_json_rows — %s: %d rows after filter %s",
                file_name,
                len(rows),
                filter_by,
            )
        return rows

    def read_excel(self, file_name: str, sheet_name: str = "Sheet1") -> list[dict]:
        """Reads an Excel file and returns a list of row dictionaries.

        The first row is treated as the header. Each subsequent row becomes
        a dictionary keyed by column headers. Empty rows are skipped.

        Args:
            file_name (str): Name of the .xlsx file.
            sheet_name (str): Name of the sheet to read. Defaults to 'Sheet1'.

        Returns:
            list[dict]: List of dictionaries, one per data row.

        Raises:
            TestDataException: If the file is missing, the sheet does not exist,
                or the file cannot be parsed.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as err:
            raise TestDataException(file_name, "openpyxl is not installed") from err

        file_path = self._get_file_path(file_name)
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as err:
            logger.error("[DATA] Failed to open Excel file %s: %s", file_name, err)
            raise TestDataException(file_name, str(err)) from err

        if sheet_name not in wb.sheetnames:
            wb.close()
            logger.error(
                "[DATA] Sheet '%s' not found in %s. Available: %s",
                sheet_name,
                file_name,
                wb.sheetnames,
            )
            raise TestDataException(file_name, f"sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            wb.close()
            logger.warning(
                "[DATA] Excel sheet '%s' in %s has no header row", sheet_name, file_name
            )
            return []

        rows = [
            dict(zip(headers, row))
            for row in rows_iter
            if any(cell is not None for cell in row)
        ]
        wb.close()

        if not rows:
            logger.warning(
                "[DATA] Excel sheet '%s' in %s has no data rows", sheet_name, file_name
            )
        else:
            logger.info(
                "[DATA] Excel loaded — %s (sheet: %s): %d rows",
                file_name,
                sheet_name,
                len(rows),
            )
        return rows
